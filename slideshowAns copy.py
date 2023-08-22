#Q1
weight = int(input("Could you Pleaze enter your weight here please: "))
x = 2.205
print(weight * x)

#Q2
import datetime
birthyear = int(input("Please enter your birth year:"))
currentYear = datetime.datetime.now().year
age = currentYear - birthyear
print("Your age is: " + str(age))

#Q3
nationality = input("Please enter your Nationality : ")
age = int(input("Please enter your Age : "))
if nationality == "Jordanian"or"jordanian" and age >= 18 :
    print("You have access to vote")
else:
        print("access not gained")

#Q4
membership_period = int(input("Please enter how many months have you been a member for: "))
age = int(input("Please enter your age: "))
cost = 150
discount = .3
costAfterDiscount = cost * discount
if membership_period >= 24 or age <= 16:
    print("Total cost = " + str(int(cost - costAfterDiscount )))
else:
        print("Total cost = " + cost)

#Q5
num = int(input("Enter grade out of 100: "))

if 90 <= num <= 100 :
        print("You Achieved an A")
elif 80 <= num <= 90:
        print("You achived a B")
elif 70 <= num <= 80 :
        print("You have achived a C")
elif 60 <= num <= 70 :
        print("You have achived a D")
elif 0 < num < 60 : 
       print("You have achived a F")
else:
       print("invalid grade")

#Q6
import random

userNum = 0

guessGame = (random.randrange(1, 6))
while userNum != guessGame:
    
    userNum = int(input("guess a number from 1-5: "))
    if guessGame != userNum:
       print("try again")
    elif guessGame == userNum:
    
        break
print("Good job")

#Q7
num = int(input("Please Enter A Number: "))

if num %2 == 0:
      print("Even Number")
else:
      print("Odd Number")

#Q8
price = int(input("Please Enter The Price: "))
voucher = input("Do You Have A Voucher (yes/no): ")

if voucher == "yes":
    voucherValue = float(input("Please Enter Voucher Value: "))
    discountValue = price * voucherValue
    priceAfterDiscount = price - discountValue
    print("Total Price is =", priceAfterDiscount)
else:
    print("Total Price is =", price)

#Q9
import random

x = 3
userguess = 0

guessGame = random.randrange(1, 101)

while userguess != guessGame:
    userguess = int(input(f"You have {x} attempts, guess the number: "))
    x -= 1

    if userguess == guessGame:
        print("Correct")
        break
    elif x == 0:
        print("GAME OVER")
        break
    elif userguess != guessGame:
        print("Try again")


#Q10
def case1():
      print("Displays  All Instructions With Description")
def case2():
      print("Control The Car")
def case3():
      print("Exits Loop")

cases = {
      "Help" : case1,
      "help" : case1,
      "start and stop" : case2,
      "Start and stop" : case2,
      "quit" : case3,
      "Quit" : case3
}
choice = input("Enter an option(Help, Start and stop, Quit): ")
if choice in (cases):
      cases[choice]()
else:
      print("invalid choice")

#Q10*
choice = input("Please enter an option(Help, Start and stop, Quit): ")
if choice == "Help"or"help":
      print("Displays  All Instructions With Description")

elif choice == "Start and stop" or "start and stop":
      print("Control The Car")

elif choice == "Quit" or "quit":
      print("Exits Loop")
else:
      print("invalid choice")

#Q10**
class car:
    def help(user):
        print("Displays  All Instructions With Description")
    def StartAndStop(user):
        print("Control The Car")
    def quit(user):
        print("Exits Loop")
car1 = car()
options = {
    "help": car1.help,
    "start and stop": car1.StartAndStop,
    "quit": car1.quit
    }
user = input("Enter an option(Help, Start and stop, Quit): ")
if user in options:
 options[user]()
else:
    print("Invalid choice")

#Q11
user = input("Enter a word/ sentence: ")
for x in user:
      print(x)

#Q12
item_list = [("banana", 44), ("milkchocolate", 37), ("chips", 12)]

total = sum(quantity for item, quantity in item_list)

print("Total Price for all items =", total)

#Q13
def my_function():
    for x in range(0,4):
        for y in range(7,10):
          print((x, y))
my_function()
#Q13*
def my_function():
    for x in range(0, 4):
        y = 7 + x 
        print((x, y))

my_function()

#Q14
number_list = []

while True:
    user = input("Enter a number (or type 'quit' to finish): ")
    
    if user == "quit":
        break
    
    number = int(user)
    number_list.append(number)

number_list.sort()

print("Length of the list:", len(number_list))
print("Sorted numbers:", number_list)

#Q15
number_list = set([2,4,6,4,3,2,3,4,6,7,3,9])
print(number_list)

#Q16
list = []

while len(list) < 5:
    user = input("Enter 5 items (or type 'quit' to finish): ")
    
    if user in list:
        print("Item already exists. Enter a new item.")
        continue
    
    list.append(user)

#Q17
def number_list(user, userPrice, userDiscount):

 return y

user = int(input("Enter Number Of Items: "))
userPrice = int(input("Enter the Price of Each Item: "))
userDiscount = int(input("Enter Discount Value: "))

x = user* ((userPrice * userDiscount)/100)
y = (userPrice * user) - x
rounded_number = round(y, 2)

print("the total price is = ", rounded_number)

#Q18
def my_function():
    num_values = int(input("Please Enter The Number Of Values: "))
    x = 0

    for _ in range(num_values):
        value = int(input("Please Enter A Value: "))
        x += value

    average = x / num_values
    return average

result = my_function()
rounded_number = round(result, 2)

print("Average:", rounded_number)

#Q19
def max_num():
     num1 = int(input("Please Enter A Number: "))
     num2 = int(input("Please Enter A Second Number: "))
     num3 = int(input("Please Enter A Third Number: "))
     
     x = max(num1, num2, num3)
     print(x)
     
max_num()

#Q20
def my_function():
    number_list = []
    total = 0
    
    while True:
        user = input("Enter a number (or type 'quit' to finish): ")
        
        if user == "quit":
            break
        total += int(user)
    
    return total

result = my_function()
print("The sum is:", result)

#Q21
def my_list():
    num_list = [2, 7, 8, 9, 4, 5, 8, 7, 4, 1]
    even_numbers = [num for num in num_list if num % 2 == 0]
    return even_numbers

E_numlist = my_list()
print("Even Numbers In The List:", E_numlist )

#Q22
def my_list():
    num_list = [2, 7, 8, 9, 4, 5, 8, 7, 4, 1]
    even_numbers = [num for num in num_list if num % 2 == 0]
    odd_numbers = [num for num in num_list if num % 2 != 0]
    return odd_numbers, even_numbers

numlist = my_list()
print("Even Numbers In The List: ", numlist[1])
print("odd Numbers In The List: ", numlist[0])

#Q23
def CheckingPrime(num):
 if num <= 1:
        return ("This Is Not A Prime Number")
 if num > 1:
      for i in range(2,num):
       if (num % i) == 0:
           
           return ("This Is Not A Prime Number")
           
       else:
        return ("This Is A Prime Number")
        break
       
user_input = int(input("Enter A Number: "))
test = CheckingPrime(user_input)

print(test)

#Q24
class Math_tools:
     def __init__(self, x, y, z):
          self.num1 = x
          self.num2= y
          self.num3 = z

class Maximum_num(Math_tools):        
     def maxNum(self):
          return max(self.num1, self.num2, self.num3)

class even_num(Math_tools):
     def evenNum(self):
               
      if x %2 == 0:
       return "Even Number"
      else:
       return "Odd Number"
      
class prime_num(Math_tools):
     def primeNum(self):

      if x <= 1:
            return ("This Is Not A Prime Number")
     
      for i in range(2, x):
       if (x % i) == 0:
           
            return "This Is Not A Prime Number"
           
       else:
            return "This Is A Prime Number"
            break
       
x = int(input("enter a number: "))
y = int(input("enter a second number: "))
z = int(input("enter a third number: "))

extractMax = Maximum_num(x, y, z)
maximum = extractMax.maxNum()
print("The maximum number is:", maximum)

extractEven = even_num(x, y, z)
even = extractEven.evenNum()
print("The given number is: ", even)

primeChecking = prime_num(x, y, z)
prime = primeChecking.primeNum()
print("The given number is: ", prime)

#Q25 
import openpyxl

excel_sheet = openpyxl.load_workbook("transcript.xlsx")

modifiedExcelsheet = excel_sheet.active

for row in modifiedExcelsheet.iter_rows(min_row=2):
    student_name = row[1]

    grades = []
    for _ in range(3):
        grade = int(input(f"Enter grade for {student_name}: "))
        grades.append(grade)

    average_grade = sum(grades) / len(grades)

    modifiedExcelsheet.cell(row=row[0].row, column=4, value=average_grade)

excel_sheet.save("transcript.xlsx")

topStudent = max(modifiedExcelsheet.iter_cols(min_row=2, max_row=modifiedExcelsheet.max_row, min_col=4, max_col=4))

top_student_name = ""
for row in modifiedExcelsheet.iter_rows(min_row=2):
    if row[4].value == topStudent:
        top_student_name = row[1].value
        break

modifiedExcelsheet.cell(row=4, column=8, value=top_student_name)

excel_sheet.save("transcript.xlsx")a




     

     
     


    






